import openpyxl

# Open the Excel workbook
wb = openpyxl.load_workbook('TestForPlotsv2.xlsx')
ws = wb.active  # Assuming the data is in the active sheet

#Building up on each Test Index:

#Test Index 1:
ws['K2'] = '= SUM(E2:G2)'
ws['K3'] = '= SUM(E3:G3)'
ws['K4'] = '= SUM(E4:G4)'
ws['K5'] = '= SUM(E5:G5)'
ws['K6'] = '= SUM(E6:G6)'
ws['K7'] = '= SUM(E7:G7)'

#Test Index 2:
ws['L2'] = '= AVERAGE(H2:J2)'
ws['L3'] = '= AVERAGE(H3:J3)'
ws['L4'] = '= AVERAGE(H4:J4)'
ws['L5'] = '= AVERAGE(H5:J5)'
ws['L6'] = '= AVERAGE(H6:J6)'
ws['L7'] = '= AVERAGE(H7:J7)'

#Test Index 3:
ws['M2'] = '= EXP(F2)'
ws['M3'] = '= EXP(F3)'
ws['M4'] = '= EXP(F4)'
ws['M5'] = '= EXP(F5)'
ws['M6'] = '= EXP(F6)'
ws['M7'] = '= EXP(F7)'

#Test Index 4:
ws['N2'] = '= -(I2)'
ws['N3'] = '= -(I3)'
ws['N4'] = '= -(I4)'
ws['N5'] = '= -(I5)'
ws['N6'] = '= -(I6)'
ws['N7'] = '= -(I7)'

#Test Index 5:
ws['O2'] = '= AVERAGE(C2:D2)'
ws['O3'] = '= AVERAGE(C3:D3)'
ws['O4'] = '= AVERAGE(C4:D4)'
ws['O5'] = '= AVERAGE(C5:D5)'
ws['O6'] = '= AVERAGE(C6:D6)'
ws['O7'] = '= AVERAGE(C7:D7)'

#Test Index 6:
ws['P2'] = '= (E2)'
ws['P3'] = '= (E3)'
ws['P4'] = '= (E4)'
ws['P5'] = '= (E5)'
ws['P6'] = '= (E6)'
ws['P7'] = '= (E7)'

#Test Index 7:
ws['Q2'] = '= (H2)'
ws['Q3'] = '= (H3)'
ws['Q4'] = '= (H4)'
ws['Q5'] = '= (H5)'
ws['Q6'] = '= (H6)'
ws['Q7'] = '= (H7)'

#Test Index 8:
ws['R2'] = '= (B2)'
ws['R3'] = '= (B3)'
ws['R4'] = '= (B4)'
ws['R5'] = '= (B5)'
ws['R6'] = '= (B6)'
ws['R7'] = '= (B7)'

#Test Index 9:
ws['S2'] = '= EXP(-J2)'
ws['S3'] = '= EXP(-J3)'
ws['S4'] = '= EXP(-J4)'
ws['S5'] = '= EXP(-J5)'
ws['S6'] = '= EXP(-J6)'
ws['S7'] = '= EXP(-J7)'

#Test Index 10:
ws['T2'] = '= (F2 * F2)'
ws['T3'] = '= (F3 * F3)'
ws['T4'] = '= (F4 * F4)'
ws['T5'] = '= (F5 * F5)'
ws['T6'] = '= (F6 * F6)'
ws['T7'] = '= (F7 * F7)'

#Test Index 11:
ws['U2'] = '= TAN(F2)'
ws['U3'] = '= TAN(F3)'
ws['U4'] = '= TAN(F4)'
ws['U5'] = '= TAN(F5)'
ws['U6'] = '= TAN(F6)'
ws['U7'] = '= TAN(F7)'

#Test Index 12:
ws['V2'] = '= AVERAGE(B2:J2)'
ws['V3'] = '= AVERAGE(B3:J3)'
ws['V4'] = '= AVERAGE(B4:J4)'
ws['V5'] = '= AVERAGE(B5:J5)'
ws['V6'] = '= AVERAGE(B6:J6)'
ws['V7'] = '= AVERAGE(B7:J7)'

#Test Index 13:
ws['W2'] = '= -(F2-G2)'
ws['W3'] = '= -(F3-G3)'
ws['W4'] = '= -(F4-G4)'
ws['W5'] = '= -(F5-G5)'
ws['W6'] = '= -(F6-G6)'
ws['W7'] = '= -(F7-G7)'

#Test Index 14:
ws['X2'] = '= (MAX(B2:D2) + MIN(B2:D2))/2'
ws['X3'] = '= (MAX(B3:D3) + MIN(B3:D3))/2'
ws['X4'] = '= (MAX(B4:D4) + MIN(B4:D4))/2'
ws['X5'] = '= (MAX(B5:D5) + MIN(B5:D5))/2'
ws['X6'] = '= (MAX(B6:D6) + MIN(B6:D6))/2'
ws['X7'] = '= (MAX(B7:D7) + MIN(B7:D7))/2'

#Test Index 15:
ws['Y2'] = '= 0.3*B2 + 0.59*C2 + 0.11*D2'
ws['Y3'] = '= 0.3*B3 + 0.59*C3 + 0.11*D3'
ws['Y4'] = '= 0.3*B4 + 0.59*C4 + 0.11*D4'
ws['Y5'] = '= 0.3*B5 + 0.59*C5 + 0.11*D5'
ws['Y6'] = '= 0.3*B6 + 0.59*C6 + 0.11*D6'
ws['Y7'] = '= 0.3*B7 + 0.59*C7 + 0.11*D7'

#Test Index 16:
ws['Z2'] = '= SQRT(0.241*POWER(B2, 2) + 0.691*POWER(C2, 2) + 0.068*POWER(D2, 2))'
ws['Z3'] = '= SQRT(0.241*POWER(B3, 2) + 0.691*POWER(C3, 2) + 0.068*POWER(D3, 2))'
ws['Z4'] = '= SQRT(0.241*POWER(B4, 2) + 0.691*POWER(C4, 2) + 0.068*POWER(D4, 2))'
ws['Z5'] = '= SQRT(0.241*POWER(B5, 2) + 0.691*POWER(C5, 2) + 0.068*POWER(D5, 2))'
ws['Z6'] = '= SQRT(0.241*POWER(B6, 2) + 0.691*POWER(C6, 2) + 0.068*POWER(D6, 2))'
ws['Z7'] = '= SQRT(0.241*POWER(B7, 2) + 0.691*POWER(C7, 2) + 0.068*POWER(D7, 2))'

#Test Index 17:
ws['AA2'] = '= TAN(G2/F2)'
ws['AA3'] = '= TAN(G3/F3)'
ws['AA4'] = '= TAN(G4/F4)'
ws['AA5'] = '= TAN(G5/F5)'
ws['AA6'] = '= TAN(G6/F6)'
ws['AA7'] = '= TAN(G7/F7)'

#Test Index 18:
ws['AB2'] = '= ATAN(G2/F2)'
ws['AB3'] = '= ATAN(G3/F3)'
ws['AB4'] = '= ATAN(G4/F4)'
ws['AB5'] = '= ATAN(G5/F5)'
ws['AB6'] = '= ATAN(G6/F6)'
ws['AB7'] = '= ATAN(G7/F7)'

#Test Index 19:
ws['AC2'] = '= SQRT(POWER(F2, 2) + POWER(G2, 2))'
ws['AC3'] = '= SQRT(POWER(F3, 2) + POWER(G3, 2))'
ws['AC4'] = '= SQRT(POWER(F4, 2) + POWER(G4, 2))'
ws['AC5'] = '= SQRT(POWER(F5, 2) + POWER(G5, 2))'
ws['AC6'] = '= SQRT(POWER(F6, 2) + POWER(G6, 2))'
ws['AC7'] = '= SQRT(POWER(F7, 2) + POWER(G7, 2))'

#Test Index 20:
ws['AD2'] = '= (MAX(B2:D2) - MIN(B2:D2))/MAX(B2:D2)'
ws['AD3'] = '= (MAX(B3:D3) - MIN(B3:D3))/MAX(B3:D3)'
ws['AD4'] = '= (MAX(B4:D4) - MIN(B4:D4))/MAX(B4:D4)'
ws['AD5'] = '= (MAX(B5:D5) - MIN(B5:D5))/MAX(B5:D5)'
ws['AD6'] = '= (MAX(B6:D6) - MIN(B6:D6))/MAX(B6:D6)'
ws['AD7'] = '= (MAX(B7:D7) - MIN(B7:D7))/MAX(B7:D7)'

wb.save('TestForPlotsv2.xlsx')