import openpyxl
from math import exp, tan, sqrt, atan

def calculate_formula(formula, values):
    """
    Evaluate the formula with given values.
    """
    # Replace None values with 0
    values = {key: value if value is not None else 0 for key, value in values.items()}

    try:
        # Include atan in the local namespace
        local_namespace = {'atan': atan}
        values.update(local_namespace)
        
        result = eval(formula, None, values)
        return result
    except Exception as e:
        print(f"Error evaluating formula '{formula}': {e}")
        return None

# Open the Excel workbook
wb = openpyxl.load_workbook('TestForPlotsv2.xlsx')
ws = wb.active  # Assuming the data is in the active sheet

# Define the formulas for each test index
formulas = {
    "TI1": "L + a + b",
    "TI2": "(H + S + I) / 3",  # Average of H, S, I
    "TI3": "exp(a)",
    "TI4": "-S",
    "TI5": "(G + B) / 2",  # Average of G, B
    "TI6": "L",
    "TI7": "H",
    "TI8": "R",
    "TI9": "exp(-I)",
    "TI10": "a ** 2",
    "TI11": "tan(a)",
    "TI12": "R + G + B + L + a + b + H + S + I",
    "TI13": "-(a - b)",
    "TI14": "(max(R, G, B) + min(R, G, B)) / 2",
    "TI15": "0.3 * R + 0.59 * G + 0.11 * B",
    "TI16": "sqrt(0.241 * R ** 2 + 0.691 * G ** 2 + 0.068 * B ** 2)",
    "TI17": "tan(b / a)",
    "TI18": "atan(b / a)",  # arctan(b / a)
    "TI19": "sqrt(a ** 2 + b ** 2)",
    "TI20": "(max(R, G, B) - min(R, G, B)) / max(R, G, B)"
}

# Iterate over rows 2 to 7 (both inclusive)
for row in range(2, 8):
    # Get the values for the current row
    values = {
        "R": ws.cell(row=row, column=3).value,
        "G": ws.cell(row=row, column=4).value,
        "B": ws.cell(row=row, column=5).value,
        "L": ws.cell(row=row, column=6).value,
        "a": ws.cell(row=row, column=7).value,
        "b": ws.cell(row=row, column=8).value,
        "H": ws.cell(row=row, column=9).value,
        "S": ws.cell(row=row, column=10).value,
        "I": ws.cell(row=row, column=11).value
    }

    # Iterate over each test index
    for col_index, (ti, formula) in enumerate(formulas.items(), start=13):
        # Calculate the value using the formula and write to the corresponding cell
        result = calculate_formula(formula, values)
        ws.cell(row=row, column=col_index, value=result)

# Save the workbook
wb.save('TestForPlotsv2_Results.xlsx')
