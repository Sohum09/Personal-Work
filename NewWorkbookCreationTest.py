import openpyxl
from openpyxl import Workbook

wb = openpyxl.load_workbook('Pesticide data_from April2021.xlsx')
ws = wb.active
last_row = ws.max_row

#Step 1. Find how many concentration values are present:
conc_count = 0
i = 2
while ws.cell(row=i, column=3).value is not None:
    conc_count += 1
    i+=1

#Step 2. Build the concentration array
conc = list()

for i in range(2, 2+conc_count):
    conc.append(ws['C{}'.format(i)].value)

#Step 3. Initialize a new workbook
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

#Step 4: Set up the first row:
#new_ws.append(['Concentration', 'R', 'G', 'B', 'L', 'a', 'b', 'H', 'S', 'I'])
new_ws['A1'] = 'Concentration'
new_ws['B1'] = 'R'
new_ws['C1'] = 'G'
new_ws['D1'] = 'B'
new_ws['E1'] = 'L'
new_ws['F1'] = 'a'
new_ws['G1'] = 'b'
new_ws['H1'] = 'H'
new_ws['I1'] = 'S'
new_ws['J1'] = 'I'

#Step 4. Time to build the data one by one.

#4A. First build the concentration:
cnt=0
for i in range(2, 2+conc_count):
    new_ws['A{}'.format(i)] = conc[cnt]
    cnt += 1

#4B. Now, we have to find the average for each data cell and input it into the new file.
for i in range(2, 2+conc_count):
    R_var = 0
    G_var = 0
    B_var = 0
    L_var = 0
    a_var = 0
    b_var = 0
    H_var = 0
    S_var = 0
    I_var = 0
    cnt = 0
    for j in range(i, last_row, 7):
        R_var += ws['D{}'.format(j)].value
        G_var += ws['E{}'.format(j)].value
        B_var += ws['F{}'.format(j)].value
        L_var += ws['G{}'.format(j)].value
        a_var += ws['H{}'.format(j)].value
        b_var += ws['I{}'.format(j)].value
        H_var += ws['J{}'.format(j)].value
        S_var += ws['K{}'.format(j)].value
        I_var += ws['L{}'.format(j)].value
        cnt += 1
    new_ws['B{}'.format(i)] = R_var / cnt
    new_ws['C{}'.format(i)] = G_var / cnt
    new_ws['D{}'.format(i)] = B_var / cnt
    new_ws['E{}'.format(i)] = L_var / cnt
    new_ws['F{}'.format(i)] = a_var / cnt
    new_ws['G{}'.format(i)] = b_var / cnt
    new_ws['H{}'.format(i)] = H_var / cnt
    new_ws['I{}'.format(i)] = S_var / cnt
    new_ws['J{}'.format(i)] = I_var / cnt

new_wb.save('OutputPlot.xlsx')