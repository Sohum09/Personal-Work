import openpyxl
import numpy as np
from scipy import stats

#Open the Excel workbook and the worksheet
wb = openpyxl.load_workbook('TestForPlots.xlsx')
ws = wb['TI']

#Initialize the concentration (in PPB) and a 2D list for all the 20 indices
conc = list()
testIndices = list()

#Building the x-value and the y-values of the lists
conc = np.array([ws.cell(row=i, column=1).value for i in range(2, 8)])
for i in range(2, 22):
    testIndex = [ws.cell(row=j, column=i).value for j in range(2, 8)]
    testIndices.append(testIndex)
testIndices = np.asarray(testIndices)

#Calculation of the equation of the trendline and the r2 value
def calcGraph(x, y):
    slope, intercept, r, p, std_err = stats.linregress(x, y)
    def func(x1):
        return slope*x1 + intercept
    modelTest = list(map(func, x))

    eqn = "y = {:.3f}x + {:.3f}".format(slope, intercept)

    residuals = y - modelTest
    ss_res = np.sum(residuals**2)
    ss_tot = np.sum((y - np.mean(y))**2)
    r_squared = 1 - (ss_res / ss_tot)

    return eqn, r_squared

eqnTI = list()
r2TI = list()

#Calculate the eqn of trend line and r2 values for all the test indices
for i in range(20):
    eqn, r2 = calcGraph(conc, testIndices[i])
    eqnTI.append(eqn)
    r2TI.append(r2)

#Display the data in tabular format
print("T.I.\t\tEquation of Trendline\t\tR2 Value")
print("-------------------------------------------------------------")
for i in range(20):
    print("{}\t\t{}\t\t{:.5f}".format(i+1, eqnTI[i], r2TI[i]))

#Finding the top 3 indices:
N = 3
res = sorted(range(len(r2TI)), key = lambda sub: r2TI[sub])[-N:]
print("\nThe top three indices in this image are:\n1. T.I. {2}\n2. T.I. {1}\n3. T.I. {0}".format(res[0]+1, res[1]+1, res[2]+1))