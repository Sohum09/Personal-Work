import tkinter as tk
from tkinter import filedialog
import openpyxl
import numpy as np
import math
from scipy import stats

def open_file_dialog():
    file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        return file_path
    else:
        return None

def load_workbook_and_process(file_path, text_widget):
    if file_path:
        # Extracting file name from the path
        file_name = file_path.split("/")[-1]  # Adjust the separator based on your operating system

        # Open the Excel workbook and the worksheet
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        #Building up on each Test Index:

        #Test Index 1:
        ws['K2'] = sum([ws['E2'].value, ws['F2'].value, ws['G2'].value])
        ws['K3'] = sum([ws['E3'].value, ws['F3'].value, ws['G3'].value])
        ws['K4'] = sum([ws['E4'].value, ws['F4'].value, ws['G4'].value])
        ws['K5'] = sum([ws['E5'].value, ws['F5'].value, ws['G5'].value])
        ws['K6'] = sum([ws['E6'].value, ws['F6'].value, ws['G6'].value])
        ws['K7'] = sum([ws['E7'].value, ws['F7'].value, ws['G7'].value])

        #Test Index 2:
        ws['L2'] = (ws['H2'].value + ws['I2'].value + ws['J2'].value)/3
        ws['L3'] = (ws['H3'].value + ws['I3'].value + ws['J3'].value)/3
        ws['L4'] = (ws['H4'].value + ws['I4'].value + ws['J4'].value)/3
        ws['L5'] = (ws['H5'].value + ws['I5'].value + ws['J5'].value)/3
        ws['L6'] = (ws['H6'].value + ws['I6'].value + ws['J6'].value)/3
        ws['L7'] = (ws['H7'].value + ws['I7'].value + ws['J7'].value)/3

        #Test Index 3:
        ws['M2'] = math.exp(ws['F2'].value)
        ws['M3'] = math.exp(ws['F3'].value)
        ws['M4'] = math.exp(ws['F4'].value)
        ws['M5'] = math.exp(ws['F5'].value)
        ws['M6'] = math.exp(ws['F6'].value)
        ws['M7'] = math.exp(ws['F7'].value)

        #Test Index 4:
        ws['N2'] = -(ws['I2'].value)
        ws['N3'] = -(ws['I3'].value)
        ws['N4'] = -(ws['I4'].value)
        ws['N5'] = -(ws['I5'].value)
        ws['N6'] = -(ws['I6'].value)
        ws['N7'] = -(ws['I7'].value)

        #Test Index 5:
        ws['O2'] = (ws['C2'].value + ws['D2'].value)/2
        ws['O3'] = (ws['C3'].value + ws['D3'].value)/2
        ws['O4'] = (ws['C4'].value + ws['D4'].value)/2
        ws['O5'] = (ws['C5'].value + ws['D5'].value)/2
        ws['O6'] = (ws['C6'].value + ws['D6'].value)/2
        ws['O7'] = (ws['C7'].value + ws['D7'].value)/2

        #Test Index 6:
        ws['P2'] = ws['E2'].value
        ws['P3'] = ws['E3'].value
        ws['P4'] = ws['E4'].value
        ws['P5'] = ws['E5'].value
        ws['P6'] = ws['E6'].value
        ws['P7'] = ws['E7'].value

        #Test Index 7:
        ws['Q2'] = ws['H2'].value
        ws['Q3'] = ws['H3'].value
        ws['Q4'] = ws['H4'].value
        ws['Q5'] = ws['H5'].value
        ws['Q6'] = ws['H6'].value
        ws['Q7'] = ws['H7'].value

        #Test Index 8:
        ws['R2'] = ws['B2'].value
        ws['R3'] = ws['B3'].value
        ws['R4'] = ws['B4'].value
        ws['R5'] = ws['B5'].value
        ws['R6'] = ws['B6'].value
        ws['R7'] = ws['B7'].value

        #Test Index 9:
        ws['S2'] = math.exp(-(ws['J2'].value))
        ws['S3'] = math.exp(-(ws['J3'].value))
        ws['S4'] = math.exp(-(ws['J4'].value))
        ws['S5'] = math.exp(-(ws['J5'].value))
        ws['S6'] = math.exp(-(ws['J6'].value))
        ws['S7'] = math.exp(-(ws['J7'].value))

        #Test Index 10:
        ws['T2'] = (ws['F2'].value) ** 2
        ws['T3'] = (ws['F3'].value) ** 2
        ws['T4'] = (ws['F4'].value) ** 2
        ws['T5'] = (ws['F5'].value) ** 2
        ws['T6'] = (ws['F6'].value) ** 2
        ws['T7'] = (ws['F7'].value) ** 2

        #Test Index 11:
        ws['U2'] = math.tan(ws['F2'].value)
        ws['U3'] = math.tan(ws['F3'].value)
        ws['U4'] = math.tan(ws['F4'].value)
        ws['U5'] = math.tan(ws['F5'].value)
        ws['U6'] = math.tan(ws['F6'].value)
        ws['U7'] = math.tan(ws['F7'].value)

        #Test Index 12:
        ws['V2'] = (ws['B2'].value + ws['C2'].value + ws['D2'].value + ws['E2'].value + ws['F2'].value + ws['G2'].value + ws['H2'].value + ws['I2'].value + ws['J2'].value)/9
        ws['V3'] = (ws['B3'].value + ws['C3'].value + ws['D3'].value + ws['E3'].value + ws['F3'].value + ws['G3'].value + ws['H3'].value + ws['I3'].value + ws['J3'].value)/9
        ws['V4'] = (ws['B4'].value + ws['C4'].value + ws['D4'].value + ws['E4'].value + ws['F4'].value + ws['G4'].value + ws['H4'].value + ws['I4'].value + ws['J4'].value)/9
        ws['V5'] = (ws['B5'].value + ws['C5'].value + ws['D5'].value + ws['E5'].value + ws['F5'].value + ws['G5'].value + ws['H5'].value + ws['I5'].value + ws['J5'].value)/9
        ws['V6'] = (ws['B6'].value + ws['C6'].value + ws['D6'].value + ws['E6'].value + ws['F6'].value + ws['G6'].value + ws['H6'].value + ws['I6'].value + ws['J6'].value)/9
        ws['V7'] = (ws['B7'].value + ws['C7'].value + ws['D7'].value + ws['E7'].value + ws['F7'].value + ws['G7'].value + ws['H7'].value + ws['I7'].value + ws['J7'].value)/9

        #Test Index 13:
        ws['W2'] = -(ws['F2'].value - ws['G2'].value)
        ws['W3'] = -(ws['F3'].value - ws['G3'].value)
        ws['W4'] = -(ws['F4'].value - ws['G4'].value)
        ws['W5'] = -(ws['F5'].value - ws['G5'].value)
        ws['W6'] = -(ws['F6'].value - ws['G6'].value)
        ws['W7'] = -(ws['F7'].value - ws['G7'].value)

        #Test Index 14:
        ws['X2'] = (max([ws['B2'].value, ws['C2'].value, ws['D2'].value]) + min([ws['B2'].value, ws['C2'].value, ws['D2'].value]))/2
        ws['X3'] = (max([ws['B3'].value, ws['C3'].value, ws['D3'].value]) + min([ws['B3'].value, ws['C3'].value, ws['D3'].value]))/2
        ws['X4'] = (max([ws['B4'].value, ws['C4'].value, ws['D4'].value]) + min([ws['B4'].value, ws['C4'].value, ws['D4'].value]))/2
        ws['X5'] = (max([ws['B5'].value, ws['C5'].value, ws['D5'].value]) + min([ws['B5'].value, ws['C5'].value, ws['D5'].value]))/2
        ws['X6'] = (max([ws['B6'].value, ws['C6'].value, ws['D6'].value]) + min([ws['B6'].value, ws['C6'].value, ws['D6'].value]))/2
        ws['X7'] = (max([ws['B7'].value, ws['C7'].value, ws['D7'].value]) + min([ws['B7'].value, ws['C7'].value, ws['D7'].value]))/2

        #Test Index 15:
        ws['Y2'] = 0.3 * ws['B2'].value + 0.59 * ws['C2'].value + 0.11 * ws['D2'].value
        ws['Y3'] = 0.3 * ws['B3'].value + 0.59 * ws['C3'].value + 0.11 * ws['D3'].value
        ws['Y4'] = 0.3 * ws['B4'].value + 0.59 * ws['C4'].value + 0.11 * ws['D4'].value
        ws['Y5'] = 0.3 * ws['B5'].value + 0.59 * ws['C5'].value + 0.11 * ws['D5'].value
        ws['Y6'] = 0.3 * ws['B6'].value + 0.59 * ws['C6'].value + 0.11 * ws['D6'].value
        ws['Y7'] = 0.3 * ws['B7'].value + 0.59 * ws['C7'].value + 0.11 * ws['D7'].value

        #Test Index 16:
        ws['Z2'] = math.sqrt(0.241 * ws['B2'].value ** 2 + 0.691 * ws['C2'].value ** 2 + ws['D2'].value ** 2)
        ws['Z3'] = math.sqrt(0.241 * ws['B3'].value ** 2 + 0.691 * ws['C3'].value ** 2 + ws['D3'].value ** 2)
        ws['Z4'] = math.sqrt(0.241 * ws['B4'].value ** 2 + 0.691 * ws['C4'].value ** 2 + ws['D4'].value ** 2)
        ws['Z5'] = math.sqrt(0.241 * ws['B5'].value ** 2 + 0.691 * ws['C5'].value ** 2 + ws['D5'].value ** 2)
        ws['Z6'] = math.sqrt(0.241 * ws['B6'].value ** 2 + 0.691 * ws['C6'].value ** 2 + ws['D6'].value ** 2)
        ws['Z7'] = math.sqrt(0.241 * ws['B7'].value ** 2 + 0.691 * ws['C7'].value ** 2 + ws['D7'].value ** 2)

        #Test Index 17:
        ws['AA2'] = math.tan(ws['G2'].value/ws['F2'].value)
        ws['AA3'] = math.tan(ws['G3'].value/ws['F3'].value)
        ws['AA4'] = math.tan(ws['G4'].value/ws['F4'].value)
        ws['AA5'] = math.tan(ws['G5'].value/ws['F5'].value)
        ws['AA6'] = math.tan(ws['G6'].value/ws['F6'].value)
        ws['AA7'] = math.tan(ws['G7'].value/ws['F7'].value)

        #Test Index 18:
        ws['AB2'] = math.atan(ws['G2'].value/ws['F2'].value)
        ws['AB3'] = math.atan(ws['G3'].value/ws['F3'].value)
        ws['AB4'] = math.atan(ws['G4'].value/ws['F4'].value)
        ws['AB5'] = math.atan(ws['G5'].value/ws['F5'].value)
        ws['AB6'] = math.atan(ws['G6'].value/ws['F6'].value)
        ws['AB7'] = math.atan(ws['G7'].value/ws['F7'].value)

        #Test Index 19:
        ws['AC2'] = math.sqrt(ws['F2'].value ** 2 + ws['G2'].value ** 2)
        ws['AC3'] = math.sqrt(ws['F3'].value ** 2 + ws['G3'].value ** 2)
        ws['AC4'] = math.sqrt(ws['F4'].value ** 2 + ws['G4'].value ** 2)
        ws['AC5'] = math.sqrt(ws['F5'].value ** 2 + ws['G5'].value ** 2)
        ws['AC6'] = math.sqrt(ws['F6'].value ** 2 + ws['G6'].value ** 2)
        ws['AC7'] = math.sqrt(ws['F7'].value ** 2 + ws['G7'].value ** 2)

        #Test Index 20:
        ws['AD2'] = (max([ws['B2'].value, ws['C2'].value, ws['D2'].value]) - min([ws['B2'].value, ws['C2'].value, ws['D2'].value])) / (max([ws['B2'].value, ws['C2'].value, ws['D2'].value]))
        ws['AD3'] = (max([ws['B3'].value, ws['C3'].value, ws['D3'].value]) - min([ws['B3'].value, ws['C3'].value, ws['D3'].value])) / (max([ws['B3'].value, ws['C3'].value, ws['D3'].value]))
        ws['AD4'] = (max([ws['B4'].value, ws['C4'].value, ws['D4'].value]) - min([ws['B4'].value, ws['C4'].value, ws['D4'].value])) / (max([ws['B4'].value, ws['C4'].value, ws['D4'].value]))
        ws['AD5'] = (max([ws['B5'].value, ws['C5'].value, ws['D5'].value]) - min([ws['B5'].value, ws['C5'].value, ws['D5'].value])) / (max([ws['B5'].value, ws['C5'].value, ws['D5'].value]))
        ws['AD6'] = (max([ws['B6'].value, ws['C6'].value, ws['D6'].value]) - min([ws['B6'].value, ws['C6'].value, ws['D6'].value])) / (max([ws['B6'].value, ws['C6'].value, ws['D6'].value]))
        ws['AD7'] = (max([ws['B7'].value, ws['C7'].value, ws['D7'].value]) - min([ws['B7'].value, ws['C7'].value, ws['D7'].value])) / (max([ws['B7'].value, ws['C7'].value, ws['D7'].value]))

        # Force recalculation of formulas
        wb.calculation_on_save = True

        # Save the workbook (this triggers recalculation)
        wb.save(file_path)

        # Open the Excel workbook and the worksheet again, this time with data=true
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Initialize the concentration (in PPB) and a 2D list for all the 20 indices
        conc = np.array([ws.cell(row=i, column=1).value for i in range(2, 8)])
        test_indices = list()

        # Building the x-value and the y-values of the lists
        for i in range(11, 31):
            test_index_values = [ws.cell(row=j, column=i).value for j in range(2, 8)]
            test_indices.append(test_index_values)
        test_indices = np.asarray(test_indices)
        
        # Calculation of the equation of the trendline and the r2 value
        def calc_graph(x, y):
            slope, intercept, r, p, std_err = stats.linregress(x, y)
            def func(x1):
                return slope*x1 + intercept
            model_test = list(map(func, x))

            eqn = "y = {:.3f}x + {:.3f}".format(slope, intercept)

            residuals = y - model_test
            ss_res = np.sum(residuals**2)
            ss_tot = np.sum((y - np.mean(y))**2)
            r_squared = 1 - (ss_res / ss_tot)

            return eqn, r_squared

        eqn_TI = list()
        r2_TI = list()

        # Calculate the eqn of the trend line and r2 values for all the test indices
        for i in range(20):
            eqn, r2 = calc_graph(conc, test_indices[i])
            eqn_TI.append(eqn)
            r2_TI.append(r2)

        # Display the data in the Text widget
        output_text = "T.I.\t\tEquation of Trendline\t\t\t\tR2 Value\n"
        output_text += "-------------------------------------------------------------\n"
        for i in range(20):
            output_text += "{}\t\t{}\t\t\t\t{:.5f}\n".format(i+1, eqn_TI[i], r2_TI[i])

        # Finding the top 3 indices:
        N = 3
        res = sorted(range(len(r2_TI)), key=lambda sub: r2_TI[sub])[-N:]
        output_text += "\nThe top three indices in this image are:\n1. T.I. {2}\n2. T.I. {1}\n3. T.I. {0}".format(res[0]+1, res[1]+1, res[2]+1)

        output_text += "\n\nWorkbook '{}' loaded successfully.".format(file_name)

        # Insert the text into the Text widget
        text_widget.insert(tk.END, output_text)

        # Slightly expand the window vertically
        window.geometry("500x550")

# Create a Tkinter window
window = tk.Tk()
window.title("Excel Data Processor")

# Function to be called when the button is clicked
def on_button_click():
    file_path = open_file_dialog()
    load_workbook_and_process(file_path, text_widget)

# Create a button to open the file dialog
button = tk.Button(window, text="Open Excel File", command=on_button_click)
button.pack(pady=20)

# Create a Text widget to display the output
text_widget = tk.Text(window, wrap=tk.WORD)
text_widget.pack(expand=True, fill=tk.BOTH)

# Run the Tkinter event loop
window.mainloop()