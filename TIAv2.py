import tkinter as tk
from tkinter import filedialog
import openpyxl
import numpy as np
from scipy import stats

def open_file_dialog():
    file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        return file_path
    else:
        return None

def load_workbook_and_process(file_path, text_widget):
    if file_path:
        # Extracting file name from the path
        file_name = file_path.split("/")[-1]  # Adjust the separator based on your operating system

        # Open the Excel workbook and the worksheet
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Initialize the concentration (in PPB) and a 2D list for all the 20 indices
        conc = np.array([ws.cell(row=i, column=1).value for i in range(2, 8)])
        test_indices = list()

        # Building the x-value and the y-values of the lists
        for i in range(2, 22):
            test_index = [ws.cell(row=j, column=i).value for j in range(2, 8)]
            test_indices.append(test_index)
        test_indices = np.asarray(test_indices)

        # Calculation of the equation of the trendline and the r2 value
        def calc_graph(x, y):
            slope, intercept, r, p, std_err = stats.linregress(x, y)
            def func(x1):
                return slope*x1 + intercept
            model_test = list(map(func, x))

            eqn = "y = {:.3f}x + {:.3f}".format(slope, intercept)

            residuals = y - model_test
            ss_res = np.sum(residuals**2)
            ss_tot = np.sum((y - np.mean(y))**2)
            r_squared = 1 - (ss_res / ss_tot)

            return eqn, r_squared

        eqn_TI = list()
        r2_TI = list()

        # Calculate the eqn of the trend line and r2 values for all the test indices
        for i in range(20):
            eqn, r2 = calc_graph(conc, test_indices[i])
            eqn_TI.append(eqn)
            r2_TI.append(r2)

        # Display the data in the Text widget
        output_text = "T.I.\t\tEquation of Trendline\t\t\t\tR2 Value\n"
        output_text += "-------------------------------------------------------------\n"
        for i in range(20):
            output_text += "{}\t\t{}\t\t\t\t{:.5f}\n".format(i+1, eqn_TI[i], r2_TI[i])

        # Finding the top 3 indices:
        N = 3
        res = sorted(range(len(r2_TI)), key=lambda sub: r2_TI[sub])[-N:]
        output_text += "\nThe top three indices in this image are:\n1. T.I. {2}\n2. T.I. {1}\n3. T.I. {0}".format(res[0]+1, res[1]+1, res[2]+1)

        output_text += "\n\nWorkbook '{}' loaded successfully.".format(file_name)

        # Insert the text into the Text widget
        text_widget.insert(tk.END, output_text)

        # Slightly expand the window vertically
        window.geometry("500x550")

# Create a Tkinter window
window = tk.Tk()
window.title("Excel Data Processor")

# Function to be called when the button is clicked
def on_button_click():
    file_path = open_file_dialog()
    load_workbook_and_process(file_path, text_widget)

# Create a button to open the file dialog
button = tk.Button(window, text="Open Excel File", command=on_button_click)
button.pack(pady=20)

# Create a Text widget to display the output
text_widget = tk.Text(window, wrap=tk.WORD)
text_widget.pack(expand=True, fill=tk.BOTH)

# Run the Tkinter event loop
window.mainloop()