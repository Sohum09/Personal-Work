import tkinter as tk
from tkinter import filedialog
import openpyxl
import numpy as np
import math
from scipy import stats
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter import Toplevel

def open_file_dialog():
    file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        return file_path
    else:
        return None

def calc_graph(x, y):
    slope, intercept, r, p, std_err = stats.linregress(x, y)
    def func(x1):
        return slope * x1 + intercept
    model_test = list(map(func, x))

    eqn = "y = {:.3f}x + {:.3f}".format(slope, intercept)

    residuals = y - model_test
    ss_res = np.sum(residuals**2)
    ss_tot = np.sum((y - np.mean(y))**2)
    r_squared = 1 - (ss_res / ss_tot)

    return eqn, r_squared

def calc_graph_plot(x, y):
    slope, intercept, r, p, std_err = stats.linregress(x, y)
    
    def func(x1):
        return slope * x1 + intercept
    
    model_test = list(map(func, x))
    residuals = y - model_test
    ss_res = np.sum(residuals**2)
    ss_tot = np.sum((y - np.mean(y))**2)
    r_squared = 1 - (ss_res / ss_tot)

    return slope, intercept, std_err, r_squared


def plot_scatter_with_trendline(x, y):
    fig, ax = plt.subplots()
    ax.scatter(x, y, label='Data')

    slope, intercept, std_err, r2 = calc_graph_plot(x, y)
    trendline = f'y = {slope:.3f}x + {intercept:.3f}'
    ax.plot(x, slope * np.array(x) + intercept, color='red', label=trendline)

    # Adding error bars
    ax.errorbar(x, y, yerr=std_err, fmt='none', color='black', capsize=5, label='Error Bars')

    ax.legend()
    ax.set_xlabel('Concentration of Pesticide (in PPB)')
    ax.set_ylabel('Test Index Value')
    return fig


def load_workbook_and_process(file_path, text_widget, plot_frame, button_frame):
    if file_path:
        file_name = file_path.split("/")[-1]

        wb1 = openpyxl.load_workbook(file_path)
        ws1 = wb1.active
        last_row = ws1.max_row

        #Step 1. Find how many concentration values are present:
        conc_count = 0
        i = 2
        while ws1.cell(row=i, column=3).value is not None:
            conc_count += 1
            i+=1

        #Step 2. Build the concentration array
        conc = list()

        for i in range(2, 2+conc_count):
            conc.append(ws1['C{}'.format(i)].value)

        #Step 3. Initialize a new workbook
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active

        #Step 4: Set up the first row:
        #new_ws.append(['Concentration', 'R', 'G', 'B', 'L', 'a', 'b', 'H', 'S', 'I'])
        new_ws['A1'] = 'Concentration'
        new_ws['B1'] = 'R'
        new_ws['C1'] = 'G'
        new_ws['D1'] = 'B'
        new_ws['E1'] = 'L'
        new_ws['F1'] = 'a'
        new_ws['G1'] = 'b'
        new_ws['H1'] = 'H'
        new_ws['I1'] = 'S'
        new_ws['J1'] = 'I'

        #Step 4. Time to build the data one by one.

        #4A. First build the concentration:
        cnt=0
        for i in range(2, 2+conc_count):
            new_ws['A{}'.format(i)] = conc[cnt]
            cnt += 1

        #4B. Now, we have to find the average for each data cell and input it into the new file.
        for i in range(2, 2+conc_count):
            R_var = 0
            G_var = 0
            B_var = 0
            L_var = 0
            a_var = 0
            b_var = 0
            H_var = 0
            S_var = 0
            I_var = 0
            cnt = 0
            step_skip = conc_count+1
            for j in range(i, last_row+1, step_skip):
                R_var += ws1['D{}'.format(j)].value
                G_var += ws1['E{}'.format(j)].value
                B_var += ws1['F{}'.format(j)].value
                L_var += ws1['G{}'.format(j)].value
                a_var += ws1['H{}'.format(j)].value
                b_var += ws1['I{}'.format(j)].value
                H_var += ws1['J{}'.format(j)].value
                S_var += ws1['K{}'.format(j)].value
                I_var += ws1['L{}'.format(j)].value
                cnt += 1
            new_ws['B{}'.format(i)] = R_var / cnt
            new_ws['C{}'.format(i)] = G_var / cnt
            new_ws['D{}'.format(i)] = B_var / cnt
            new_ws['E{}'.format(i)] = L_var / cnt
            new_ws['F{}'.format(i)] = a_var / cnt
            new_ws['G{}'.format(i)] = b_var / cnt
            new_ws['H{}'.format(i)] = H_var / cnt
            new_ws['I{}'.format(i)] = S_var / cnt
            new_ws['J{}'.format(i)] = I_var / cnt

        new_wb.save('OutputPlot.xlsx')

        wb = openpyxl.load_workbook('OutputPlot.xlsx')
        ws = wb.active

        last_row = ws.max_row

        # Building up on each Test Index:
        for i in range(2, last_row+1):
            # Test Index 1:
            ws['K{}'.format(i)] = sum([ws['E{}'.format(i)].value, ws['F{}'.format(i)].value, ws['G{}'.format(i)].value])

            # Test Index 2:
            ws['L{}'.format(i)] = (ws['H{}'.format(i)].value + ws['I{}'.format(i)].value + ws['J{}'.format(i)].value) / 3

            # Test Index 3:
            ws['M{}'.format(i)] = math.exp(ws['F{}'.format(i)].value)

            # Test Index 4:
            ws['N{}'.format(i)] = -(ws['I{}'.format(i)].value)

            # Test Index 5:
            ws['O{}'.format(i)] = (ws['C{}'.format(i)].value + ws['D{}'.format(i)].value) / 2

            # Test Index 6:
            ws['P{}'.format(i)] = ws['E{}'.format(i)].value

            # Test Index 7:
            ws['Q{}'.format(i)] = ws['H{}'.format(i)].value

            # Test Index 8:
            ws['R{}'.format(i)] = ws['B{}'.format(i)].value

            # Test Index 9:
            ws['S{}'.format(i)] = math.exp(-(ws['J{}'.format(i)].value))

            # Test Index 10:
            ws['T{}'.format(i)] = (ws['F{}'.format(i)].value) ** 2

            # Test Index 11:
            ws['U{}'.format(i)] = math.tan(ws['F{}'.format(i)].value)

            # Test Index 12:
            ws['V{}'.format(i)] = (ws['B{}'.format(i)].value + ws['C{}'.format(i)].value + ws['D{}'.format(i)].value + ws['E{}'.format(i)].value + ws['F{}'.format(i)].value + ws['G{}'.format(i)].value + ws['H{}'.format(i)].value + ws['I{}'.format(i)].value + ws['J{}'.format(i)].value) / 9

            # Test Index 13:
            ws['W{}'.format(i)] = -(ws['F{}'.format(i)].value - ws['G{}'.format(i)].value)

            # Test Index 14:
            ws['X{}'.format(i)] = (max([ws['B{}'.format(i)].value, ws['C{}'.format(i)].value, ws['D{}'.format(i)].value]) + min([ws['B{}'.format(i)].value, ws['C{}'.format(i)].value, ws['D{}'.format(i)].value])) / 2

            # Test Index 15:
            ws['Y{}'.format(i)] = 0.3 * ws['B{}'.format(i)].value + 0.59 * ws['C{}'.format(i)].value + 0.11 * ws['D{}'.format(i)].value

            # Test Index 16:
            ws['Z{}'.format(i)] = math.sqrt(0.241 * ws['B{}'.format(i)].value ** 2 + 0.691 * ws['C{}'.format(i)].value ** 2 + ws['D{}'.format(i)].value ** 2)

            # Test Index 17:
            ws['AA{}'.format(i)] = math.tan(ws['G{}'.format(i)].value / ws['F{}'.format(i)].value)

            # Test Index 18:
            ws['AB{}'.format(i)] = math.atan(ws['G{}'.format(i)].value / ws['F{}'.format(i)].value)

            # Test Index 19:
            ws['AC{}'.format(i)] = math.sqrt(ws['F{}'.format(i)].value ** 2 + ws['G{}'.format(i)].value ** 2)

            # Test Index 20:
            ws['AD{}'.format(i)] = (max([ws['B{}'.format(i)].value, ws['C{}'.format(i)].value, ws['D{}'.format(i)].value]) -
                                 min([ws['B{}'.format(i)].value, ws['C{}'.format(i)].value, ws['D{}'.format(i)].value])) / \
                                (max([ws['B{}'.format(i)].value, ws['C{}'.format(i)].value, ws['D{}'.format(i)].value]))

        conc = np.array([ws.cell(row=i, column=1).value for i in range(2, last_row+1)])
        test_indices = list()

        for i in range(11, 31):
            test_index_values = [ws.cell(row=j, column=i).value for j in range(2, last_row+1)]
            test_indices.append(test_index_values)
        test_indices = np.asarray(test_indices)

        output_text = "T.I.\t\tEquation of Trendline\t\t\t\tR2 Value\n"
        output_text += "-------------------------------------------------------------\n"

        eqn_TI = list()
        r2_TI = list()
            
        for i in range(20):
            eqn, r2 = calc_graph(conc, test_indices[i])
            output_text += "{}\t\t{}\t\t\t\t{:.5f}\n".format(i+1, eqn, r2)
            eqn_TI.append(eqn)
            r2_TI.append(r2)
            button = tk.Button(button_frame, text=f"Plot T.I. {i + 1}", command=lambda i=i: plot_button_click(i, conc, test_indices[i]))
            button.pack(side=tk.TOP) # Placing the buttons alongside the output data.

        # Finding the top 3 indices:
        N = 3
        res = sorted(range(len(r2_TI)), key=lambda sub: r2_TI[sub])[-N:]
        output_text += "\nThe top three indices in this image are:\n1. T.I. {2}\n2. T.I. {1}\n3. T.I. {0}".format(res[0]+1, res[1]+1, res[2]+1)
        
        # Acknowledge succesful loading of the application.
        output_text += "\n\nWorkbook '{}' loaded successfully.".format(file_name)

        text_widget.insert(tk.END, output_text)

def plot_button_click(index, x, y):
    new_window = Toplevel(window)
    new_window.title(f"Plot for T.I. {index + 1}")

    slope, intercept, std_err, r2 = calc_graph_plot(x, y)
    
    fig = plot_scatter_with_trendline(x, y)
    canvas = FigureCanvasTkAgg(fig, master=new_window)
    canvas.draw()

    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

    def on_close_fig():
        plt.close(fig)
        new_window.destroy()

    new_window.protocol("WM_DELETE_WINDOW", on_close_fig)

if __name__ == "__main__":
    # Create a Tkinter window
    window = tk.Tk()
    window.title("Excel Data Processor")

    def on_button_click():
        file_path = open_file_dialog()
        load_workbook_and_process(file_path, text_widget, plot_frame, button_frame)

    def on_closing():
        window.destroy()

    window.protocol("WM_DELETE_WINDOW", on_closing)

    button = tk.Button(window, text="Open Excel File", command=on_button_click)
    button.pack(pady=20)

    button_frame = tk.Frame(window)
    button_frame.pack(side=tk.LEFT, fill=tk.Y)

    text_widget = tk.Text(window, wrap=tk.WORD)
    text_widget.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

    plot_frame = tk.Frame(window)
    plot_frame.pack(expand=True, fill=tk.BOTH)

    window.mainloop()   
