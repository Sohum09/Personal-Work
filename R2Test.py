import openpyxl 
import matplotlib.pyplot as plt
import numpy as np
from scipy import stats

wb = openpyxl.load_workbook('SuccessfulExperiment_RSquare.xlsx')
ws = wb['R2Val']

#Inputting a range of data:
imageTest = list()
R2Data = list()

imageTest = np.array([ws.cell(row=i, column=1).value for i in range(2, 36)])
R2Data = np.array([ws.cell(row=i, column=2).value for i in range(2, 36)])

#Calculating equation of the trend line:
slope, intercept, r, p, std_err = stats.linregress(imageTest, R2Data)
print("Equation of trend line: y = {:.3f}x + {:.3f}".format(slope, intercept))

#Function to find equation of line:
def func(x):
    return slope*x + intercept
modelTest = list(map(func, imageTest))

# Calculate R-squared value and show it:
residuals = R2Data - modelTest
ss_res = np.sum(residuals**2)
ss_tot = np.sum((R2Data - np.mean(R2Data))**2)
r_squared = 1 - (ss_res / ss_tot)

print("R^2 value: {:.4f}".format(r_squared))

#Plot and show the trend line:
plt.scatter(imageTest, R2Data)
plt.plot(imageTest, modelTest)
plt.title("Just a test :)")
plt.xlabel("Image Test No.")
plt.ylabel("R^2 Score")
plt.show()